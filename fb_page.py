import requests
from bs4 import BeautifulSoup
import time
import urlparse
from openpyxl import load_workbook
from openpyxl import Workbook

def getting_industry_link():
    file_obj = open('mainpage.txt','r')
    text = file_obj.read()
    file_obj.close()
    y=[]  # to temporary store list of indusrty urls
    soup = BeautifulSoup(text,'html.parser')
    for a in soup.findAll('div'):
        for b in a.findAll('ul',{'class':'multi-dropdown-list'}):
            for c in b.findAll('li'):
                for d in c.findAll ('a'):
                    y.append(d.get('href'))
    z=set(y)
    for x in z:
        ind_url.append(x)
def changing_url(country,industry): # this function will convert the urls in a crawlerable format
    temp_list = list(industry)
    x = "http://www/socialbakers.com"+industry
    # print u1.lower()+"gggggggggg"
    # y = urlparse.urljoin(u1, industry);
    # temp_list[0:0]="http://www.socialbakers.com"
    # temp_list[38:55] = "/facebook/pages/"
    # x = urlparse()
    # x="".join(temp_list)
    print x
    return (x.lower())
def writing_to_file(name , fb_url,country,industry, localFans, totalFans, link):
    global h_num , c_num ,ind_num ,nam_num
    wb = load_workbook('facebook_india.xlsx')
    ws = wb.active
    ws.cell(row = h_num, column = 7).value=fb_url
    ws.cell(row = c_num, column = 3).value=country
    ws.cell(row = ind_num, column = 4).value=industry
    ws.cell(row = nam_num, column = 1).value=name
    ws.cell(row = lfan_num, column = 5).value=local_fans
    ws.cell(row = tfan_num, column = 6).value=total_fans
    ws.cell(row = link_num, column = 2).value=link
    h_num+=1
    c_num+=1
    ind_num+=1
    nam_num+=1
    lfan_num+=1
    tfan_num+=1
    link_num+=1
    wb.save('facebook_india.xlsx')
def page_empty_check(soup):  # this function will check if the page is empty or not
    count=0
    for a in soup.findAll('div',{'class':'brand-table-placeholder'}):  # these loops check if the page is empty or not
        for b in a.findAll('table',{'class':'brand-table-list'}):
            for c in b.findAll('tr'):
                count+=1
                if (count == 2):
                   return count
    return count
def wrong_page_check (soup): # this function checks for landing in the wrong page i.e. Whopsi error 404 page not found one
    for a in soup.findAll('title'):
        if "404" in a.string :
            return 1
        else :
            return 0
def check_show_more_button(soup): # this check if show more button is present in the page or not
    for a in soup.findAll('div',{'class':'more-center-link'}):
         for b in a.findAll('a'):
             if "Show More" in b.string :
                return 1
    return 0
def getting_facebook_page_url(url):  # this function will goto each page and get the url of the facebook page
    while True :
        try :
            sourcecode_page = requests.get(url)
            print sourcecode_page;
            break
        except requests.exceptions.ConnectionError:  # to handle error if website refuses connection due to multiple retries
             print("Wait !! Website Refusing Connection.... ")
             time.sleep(30)
    text_source_page = sourcecode_page.text
    soup_page=BeautifulSoup(text_source_page,'html.parser')
    for a in soup_page.findAll('a',{'class':'blank show-tooltip'}):
        return a.get('href')
def main(temp_url,country,ind):
    page_num_start=1
    page_num_end = 5
    while True :
        url = temp_url+"page-"+str(page_num_start)+"-"+str(page_num_end)+"/"
        print url+"llllllllllllllllllll"
        while True :
            try :
                #time.sleep(10)
                source_code=requests.get(url)  #getting page source
                # print source_code
                print source_code.url+"ffffffffffffffffffffffffffffff"
                break
            except requests.exceptions.ConnectionError:  # to handle error if website refuses connection due to multiple retries
                print("Wait !! Website Refusing Connection.... ")
                time.sleep(30)
        if country.lower() not in source_code.url :   # this checks if the url has been redirected to another webpage without country or not and if yes then it stops the loop
            print "somthing happening here"
            break

        text_source=source_code.text #converting it to text file
        print text_source+"dddddddddddddd"
        soup=BeautifulSoup(text_source,'html.parser')  # making soup object
        flag = wrong_page_check(soup)
        if flag == 1 :
            break
        count = page_empty_check(soup)
        print "ccc"
        print count
        print "ccc"
        if (count != 1) :   # if this condition is true this implies everything is fine and now getting the handles
            print "am coming here"
            b = soup.findAll('div',{'class','item'});
            print b
            for a in soup.findAll('div',{'class','item'}):
                print a+"-----------"
                print "coming here"
                for fa in a.findAll('strong'):
                    lf = fa[0].text
                    tf = fa[1].text  
                    print fa[0].text
                    print fa[1].text
                for xa in a.findAll('a'):
                    if xa.get('title') is not None :
                        link = "http://www.socialbakers.com"+xa.get('href')
                        fb_page_url = getting_facebook_page_url("http://www.socialbakers.com"+xa.get('href'))
                        print "am here man.........."+fb_page_url
                for b in a.findAll('h2'):
                    for c in b.findAll('span'):
                        name=c.string
                        #print(name+" - "+fb_page_url + " - " + country[:-1] + " - " + ind[29:-1])
                        writing_to_file(name , fb_page_url,country[:-1],ind[29:-1],lf,tf,link)
        flag2=check_show_more_button(soup)
        if flag2 == 1 :
            pass
        else :
            break
        page_num_start+=5
        page_num_end += 5

country_name=[]
ind_url=[]
h_num=2
c_num=2
ind_num=2
nam_num = 2
lfan_num=2
tfan_num=2
link_num=2

getting_industry_link()

wb = Workbook() # creating excel file
ws = wb.active
# making headers for excel file
ws.cell(row =1, column = 1).value="Name"
ws.cell(row =1, column = 2).value="Link"
ws.cell(row =1, column = 3).value="Country"
ws.cell(row =1, column = 4).value="Industry"
ws.cell(row =1, column = 5).value="Local Fans"
ws.cell(row =1, column = 6).value="Total Fans"
ws.cell(row =1, column = 7).value="Facebook URL"

wb.save('facebook_india.xlsx') # saving the file

for temp_industry_url in ind_url :
        print temp_industry_url + "this is tttttttttttt"
        url = changing_url("India/",temp_industry_url)
        print url + "uuuuuuuuuuuuuu"
        main(url,"india/",temp_industry_url)
